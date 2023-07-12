from collections import defaultdict
from os import path
import json
import pathlib
import re
import string
import mlflow
import numpy as np
from pypdf import PdfReader

from tqdm import tqdm
import spacy
import pdfplumber
from bs4 import BeautifulSoup
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font

from downloads import get_fn_of_doi, BASE_DATA_DIR, DATASET_FN, DATA_PAPERS_DIR, RESEARCH_PAPERS_DIR

FORCE_REBUILD_ENTITIES = not True

def get_spacy_model(model_name):
    try:
        spacy.load(model_name)
    except OSError:
        print(f'Model {model_name} not found. Downloading...')
        spacy.cli.download(model_name)
        print(f'Model {model_name} downloaded.')

    return spacy.load(model_name)


def save_entities_to_file(entities, output_file):
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(entities, f, ensure_ascii=False, indent=4)


def load_entities_from_file(output_file):
    with open(output_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_text_pdf(fn, engin='pypdf'):
    assert engin in ['pdfplumber', 'pypdf']
    print(fn)
    text = ''
    if engin == 'pdfplumber':
        with pdfplumber.open(fn) as pdf:
            for page in pdf.pages:
                text += page.extract_text() + '\n'
    elif engin == 'pypdf':
        reader = PdfReader(fn)
        for i in range(len(reader.pages)):
            page = reader.pages[i]
            text += page.extract_text() + '\n'
    return text

def extract_text_html(fn):
    with open(fn, "r", encoding='utf-8') as f:
        html_content = f.read()
    soup = BeautifulSoup(html_content, "html.parser")
    return soup.get_text()


def extract_text_txt(fn):
    with open(fn, "r", encoding='utf-8') as f:
        return f.read()
    

def extract_and_save_entities(df, paper_type, nlp, remove_nums, text_type):
    """Extract and save named entities of a all papers with a type in the df.
    
    Args:
        df: Linking dataset as a DataFrame.
        paper_type: Type of papers ('data' or 'research') which their entities should be extracted.
        nlp: A spaCy model.
        remove_nums: A Boolean flag indicating whether numerical entities should be removed from the resulting list.
        text_type: Either 'full-text' for article's full text or 'abstract' for article abstracts.

    """
    assert paper_type in ['data', 'research']
    assert text_type in ['full-text', 'abstract']
    entities_fn = path.join(BASE_DATA_DIR, f'{paper_type}_papers_entities_remove_nums={remove_nums}_text_type={text_type}.json')
    if path.exists(entities_fn) and not FORCE_REBUILD_ENTITIES:
        return load_entities_from_file(entities_fn)

    entities = dict()
    entities_nums_removed = dict()
    print(f'Extracting entities of {paper_type} papers ...')
    for _, row in tqdm(list(df.iterrows())):
        doi = row[f'{paper_type}_paper_doi']
        fn = get_fn_of_doi(doi, DATA_PAPERS_DIR if paper_type == 'data' else RESEARCH_PAPERS_DIR, already_exists=True)
        
        rest, ext = fn.rsplit('.', 1)
        if text_type == 'abstract':
            ext = 'txt'
            fn = rest + '.abstract.txt'

        if ext == 'pdf':
            text = extract_text_pdf(fn)
        elif ext == 'html':
            text = extract_text_html(fn)
        elif ext == 'txt':
            text = extract_text_txt(fn)
        else:
            print(f'Unsupported file format: {ext}')
            continue

        entities[doi], entities_nums_removed[doi] = extract_entities(nlp, text)

    entities_fn = path.join(BASE_DATA_DIR, f'{paper_type}_papers_entities_remove_nums=False_text_type={text_type}.json')
    entities_fn_nums_removed = path.join(BASE_DATA_DIR, f'{paper_type}_papers_entities_remove_nums=True_text_type={text_type}.json')
    save_entities_to_file(entities, entities_fn)
    save_entities_to_file(entities_nums_removed, entities_fn)
    return entities

def extract_entities(nlp, text, include_ent_types=True):
    """Extract named entities from a piece of text.
    
    Args:
        nlp: A spaCy model.
        text: The input text from which named entities will be extracted.
        remove_nums: A Boolean flag indicating whether numerical entities should be removed from the resulting list.
    """
    ents = [(e.text, e.label_) for e in nlp(text).ents]
    with open('ents_explore.txt', 'a', encoding='utf-8') as f:
        for x,y in ents:
            f.write(f'{x}:{y}\n')
    
    ret1 = [e[0] + ':' + e[1] if include_ent_types else e[0] for e in ents]
    ret1 = sorted(list(set(ret1)))
    ret2 = [e[0] + ':' + e[1] if include_ent_types else e[0] for e in ents if 
            e[1] not in ['ORDINAL', 'CARDINAL', 'PERCENT', 'DATE', 'TIME', 'MONEY'] and not e[0].isdigit()]
    ret2 = sorted(list(set(ret2)))
    return ret1, ret2

def link_a_paper_by_ner(src_doi, src_type, entities_dict):
    """Rank corresponding papers of a given paper by NER method.
    
    Args:
        src_doi: DOI of the paper that its corresponding papers are found.
        src_type: Type of the paper ('data' or 'research') that its corresponding papers should be found. 
        entities_dict: A dictionary with keys=('data' or 'research'). Each value is a dictionary as returned by extract_and_save_entities.
    """
    assert src_type in ['data', 'research']
    other_type =  'data' if src_type == 'research' else 'research'
    ranks = dict()
    src_entities = entities_dict[src_type][src_doi]
    for other_doi, other_entities in entities_dict[other_type].items():
        common_entities = set(src_entities) & set(other_entities)
        ranks[other_doi] = len(common_entities)
    return dict(sorted(ranks.items(), key=lambda item: item[1], reverse=True))


def ner_method_whole_dataset(src_type, src_dois, entities_dict, df, mlflow_on, remove_nums, text_type):
    """Apply NER method to the whole dataset. 
    
    Args:
        df: Linking dataset as a DataFrame.
        src_type: Type of source papers ('data' or 'research') that their corresponding papers are found.
    """
    assert src_type in ['data', 'research']
    other_type =  'data' if src_type == 'research' else 'research'
    correct_ranks = []
    for src_doi in src_dois:
        rank = link_a_paper_by_ner(src_doi, src_type, entities_dict)
        rank_dois = list(rank.keys())
        ans_doi = df[df[f'{src_type}_paper_doi']==src_doi].iloc[0][f'{other_type}_paper_doi']
        correct_ranks.append(rank_dois.index(ans_doi))
    
    correct_ranks = np.array(correct_ranks) + 1
    mrr = (1/correct_ranks).mean()
    rr_std = (1/correct_ranks).std()
    print(f'Finding corresponding papers of {src_type} papers')
    print(f'Reciprocal Rank: MMR={mrr:.5f}, STD={rr_std:.2f}')
    fig = plt.figure()
    ax = fig.gca()
    labels, counts = np.unique(correct_ranks, return_counts=True)
    ax.bar(labels, counts, align='center')
    ax.set_xticks(labels)
    ax.set_title(f'Histogram of correct-answer-rank for\nfinding corresponding papers of {src_type} papers (NER)')
    
    if mlflow_on:
        with mlflow.start_run():
            mlflow.log_param('method', 'NER')
            mlflow.log_param('remove_nums', remove_nums)
            mlflow.log_param('text_type', text_type)
            mlflow.log_param('src_type', src_type)
            mlflow.log_metric('mmr', mrr)
            mlflow.log_metric('rr_std', rr_std)
            mlflow.log_figure(fig, 'hist-of-ranks.png')

    return mrr, rr_std

####################################################
MLFLOW_ON = not True
SPACY_MODEL_NAME = 'en_core_web_sm'
CV_NUM_FOLDS = 5
nlp = get_spacy_model(SPACY_MODEL_NAME)

df = pd.read_csv(DATASET_FN, index_col=0)
print('Total number of pairs of data/research papers in the dataset:', len(df))

df = df[df['data_paper_has_file'] & df['research_paper_has_file']]
print('Number of pairs with both files present:', len(df))

def ner_method_all_combinations_with_cross_validation():
    """
    Note: the output of this function was used in a table in the presentation to Korean team.

    Note on the cross-validation like method employed:
    text_type and remove_nums are hyperparameters. As they have no parameters, the algorithms
    do not necessitate a conventional "training phase" but requires an optimal selection of these hyperparameters.
    Given the limited size of the dataset, a method analogous to cross-validation is employed. For
    each possible combination of hyperparameters, the algorithm is tested on K-1 folds of the data,
    This procedure, though reminiscent of cross-validation, essentially serves as a
    robust performance evaluation mechanism for each hyperparameter configuration.
    The set of hyperparameters that yield the highest average performance across all folds is then selected as the optimal choice."""
    results = defaultdict(dict)
    for text_type in ['full-text']:
        for remove_nums in [False, True]:
            print(f'\n\n####### {text_type} {remove_nums} ##########')
            entities_dict = dict()
            for paper_type in ['research', 'data']:
                entities_dict[paper_type] = extract_and_save_entities(df, paper_type, nlp, remove_nums, text_type)
            if MLFLOW_ON:
                mlflow.create_experiment('ner-all')
                mlflow.set_experiment('raw-ner')
            for src_type in ['data', 'research']:
                all_src_dois = list(entities_dict[src_type])
                CV_LEN = len(all_src_dois) // CV_NUM_FOLDS
                mrrs, rr_stds = [], []
                for k in range(CV_NUM_FOLDS):
                    if k == CV_NUM_FOLDS - 1:
                        src_dois = all_src_dois[k*CV_LEN:]
                    else:
                        src_dois = all_src_dois[k*CV_LEN:(k+1)*CV_LEN]
                    mrr, rr_std = ner_method_whole_dataset(src_type, src_dois, entities_dict, df, MLFLOW_ON, remove_nums, text_type)
                    mrrs.append(mrr)
                    rr_stds.append(rr_std)
                results[(text_type, remove_nums)][src_type] = (np.array(mrrs).mean(), np.array(rr_stds).mean())

    print(results)
    plt.show()


def ners_side_by_side():
    # global df
    # df = df[:1]
    for text_type in ['full-text']:
        for remove_nums in [False, True]:
            print(f'\n\n####### {text_type} {remove_nums} ##########')
            for paper_type in ['research', 'data']:
                extract_and_save_entities(df, paper_type, nlp, remove_nums, text_type)
    
    wb_fn = path.join(BASE_DATA_DIR, 'ners.xlsx')
    wb = openpyxl.Workbook()
    del wb['Sheet']
    for text_type in ['full-text']:
        for remove_nums in [False, True]:
            # sh_name = f'remove_nums={remove_nums}_text_type={text_type}'
            ws_name = f'{remove_nums}_{text_type}'
            wb.create_sheet(ws_name)
            ws = wb[ws_name]
            entities_dict = dict()
            for paper_type in ['research', 'data']:
                entities_dict[paper_type] = extract_and_save_entities(df, paper_type, nlp, remove_nums, text_type)

            col = 1
            row_nums = []
            for _, row in tqdm(list(df.iterrows())):
                dp_doi, rp_doi = row['data_paper_doi'], row['research_paper_doi']
                dp_nes, rp_nes = entities_dict['data'][dp_doi], entities_dict['research'][rp_doi]
                ws.cell(1,col).value = 'Data paper'
                ws.cell(1,col+1).value = 'Research paper'
                ws.cell(2,col).value = dp_doi
                dp_fn = path.abspath(get_fn_of_doi(dp_doi, DATA_PAPERS_DIR, already_exists=True))
                ws.cell(2,col).hyperlink = pathlib.Path(dp_fn).as_uri()
                ws.cell(2,col+1).value = rp_doi
                rp_fn = path.abspath(get_fn_of_doi(rp_doi, RESEARCH_PAPERS_DIR, already_exists=True))
                ws.cell(2,col+1).hyperlink = pathlib.Path(rp_fn).as_uri()
                for i, ne in enumerate(dp_nes):
                    ne = ''.join(filter(lambda x: x in string.printable, ne))
                    ws.cell(i+3,col).value = ne
                for i, ne in enumerate(rp_nes):
                    ne = ''.join(filter(lambda x: x in string.printable, ne))
                    ws.cell(i+3,col+1).value = ne
                row_nums.append(max(len(dp_nes), len(rp_nes)))
                col += 2

            fills = [PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid"),
                PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")]
            for col in range(1, col):
               ws.cell(1,col).font = Font(bold=True)
               ws.cell(2,col).font = Font(bold=True)
               ws.cell(2,col).style = 'Hyperlink'
               for row in range(1, max(row_nums)+3):
                    ws.cell(row,col).fill = fills[((col-1)//2)%2]
            
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(wb_fn)


ner_method_all_combinations_with_cross_validation()
# ners_side_by_side()